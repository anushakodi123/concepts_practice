from openpyxl import Workbook, load_workbook
from tempfile import NamedTemporaryFile
from openpyxl.styles import Font,Color, Side, Border, PatternFill, GradientFill, Alignment
from openpyxl.chart import BarChart, Series, Reference
from datetime import datetime
wb = Workbook()

ws = wb.active

ws['A1'] = 45

wb.save("sample.xlsx")

ws1 = wb.create_sheet("worksheet1")

ws.title = "changed"
print(wb.sheetnames)

target = wb.copy_worksheet(ws)
print(wb.sheetnames)
print(ws['A1'].value)

d = ws.cell(row=1, column=2, value=10)
print(ws['A1':'B1'])

colC=ws['C']
row10=ws[10]

for row in ws.iter_rows(min_row=1, max_col=2, max_row=2):
    for cell in row:
        print(cell)


for col in ws.iter_cols(min_row=1, max_col=2, max_row=2):
    for cell in col:
        print(cell)

ws = wb.active
ws['C9'] = "hello world"
print(tuple(ws.rows))


print(tuple(ws.columns))


for row in ws.values:
    for value in row:
        print(value)

for row in ws.iter_rows(min_row=1, min_col=2, max_row=2, values_only=True):
    print(row)

wb = load_workbook("sample.xlsx")
wb.template = True
wb.save("sample.xltx")


with NamedTemporaryFile() as f:
    wb.save(f.name)
    f.seek(0)
    stream =f.read()

# wb = Workbook()
# wb = load_workbook(filename="sample.xlsx")
# sheet_ranges = wb['range names']
# print(sheet_ranges['D18'].value)


wb = Workbook()
ws = wb.active
treeData = [["Type", "Leaf Color", "Height"], ["Maple", "Red", 549], ["Oak", "Green", 783], ["Pine", "Green", 1204]]

for row in treeData:
    ws.append(row)

ft = Font(bold=True)
for row in ws["A1:C1"]:
    for cell in row:
        cell.font = ft

chart = BarChart()
chart.type = "col"
chart.title = "Tree Height"
chart.y_axis.title = 'Height (cm)'
chart.x_axis.title = 'Tree Type'
chart.legend = None

data = Reference(ws, min_col=3, min_row=2, max_row=4, max_col=3)
categories = Reference(ws, min_col=1, min_row=2, max_row=4, max_col=1)


chart.add_data(data)
chart.set_categories(categories)


ws.add_chart(chart, "E1")
wb.save("TreeData.xlsx")


# wb = load_workbook("sample.xlsx")
ws = wb.active

a1 = ws['A1']

a1.font = Font(color="FF0000")
print(ws)
ws['A2'] = 42
a1.font = Font(color="FF0000")
a2 = ws['A2']
# wb.save("sample.xlsx")
a2.font = Font(bold=True,color="FF0000",size=14)
ws.merge_cells('B2:F4')
thin=Side(border_style="thin", color="000000")
double=Side(border_style="thin", color="FF0000")
a1.border = Border(top=thin, left=double, right=thin, bottom=double)
a1.fill = PatternFill(patternType="solid", fgColor="DDDDDD")
a2.fill = fill = GradientFill(stop=("000000", "FFFFFF"))
a3=ws['A3']
a3.value = "hello world"
a4=ws['A4']
a4.number_format = 'YYYY-MM-DD'
a4.alignment = Alignment(horizontal="center", vertical="center")
# a3 = datetime(2010, 7, 21)
# print(a3.number_format)
ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
ws.page_setup.paperSize = ws.PAPERSIZE_TABLOID
ws.page_setup.fitToHeight = 0
ws.page_setup.fitToWidth = 1
wb.save("sample.xlsx")